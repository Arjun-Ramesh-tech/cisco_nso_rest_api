from datetime import datetime
from multiprocessing import pool
import os
import json
import requests
from openpyxl import Workbook,load_workbook
from dotenv import load_dotenv
import time
import xml.etree.ElementTree as ET
import aiohttp
import asyncio
import sys
load_dotenv()

now = datetime.now()
dt_string = now.strftime("_%Y%m%d%H%M").strip()


class discrepancy_check:
    def __init__(self):
        self.excel_file = "NSO_discrepancy_check" + dt_string + ".xlsx"
        self.url = os.getenv(
            "NSO_URL") + "/api/config/resource-db/_operations/discrepancy-check"
        self.discrepancy_header = {}
        self.discrepancy_header["Accept"] = 'application/vnd.yang.data+json'
        self.data = []
        self.local_list_xml = os.getenv("local_pool_list_xml")
        self.global_list_xml = os.getenv("global_pool_list_xml")
        self.evi = "evi_id_l2_eline_evpn_"
        self.bdid = "bd_id_l2_elan_etree_evpn_"
        self.evpl = "evpl_id_l2_eline_evpn_"
        self.range_json = {}
        self.summary_discrepancy = []

    def create_summary_data(self):
        total_true = 0
        total_false = 0
        for each in self.response:
            count_true = 0
            count_false = 0
            for each_i in each["discrepancies"]:
                if each_i["exist-in-resource-db"] == True:
                    count_true = count_true + 1
                else:
                    count_false = count_false + 1
            total_true = total_true + count_true
            total_false = total_false + count_false
            self.summary_discrepancy.append(
                [each['pool'], count_true, count_false])
        self.summary_discrepancy.append(["Total", total_true, total_false])

    def get_localrange(self, pool_name, value):
        self.localurl_range = os.getenv(
            "NSO_URL") + "/api/running/tnt-resource-db:resource-db/local-id-pool/"
        self.localheader = {}
        self.localheader["Accept"] = 'application/vnd.yang.collection+json'
        self.localheader['Content-Type'] = 'text/plain'
        try:
            item = requests.request("GET", self.localurl_range+pool_name.rsplit("_", 1)[0]+","+pool_name.rsplit(
                "_", 1)[1]+"/range", headers=self.localheader, auth=(os.getenv("NSO_USERNAME"), os.getenv("NSO_PASSWORD")))
            if item.status_code == 200:
                item = item.json()
                self.range_json[pool_name] = item["collection"]["tnt-resource-db:range"]
                for each in self.range_json[pool_name]:
                    if (value >= each["min"] and value <= each["max"]):
                        return True
                return False
            else:
                #print("Range Call Failed")
                return "Pool range doesn't exist"
        except requests.exceptions.InvalidURL as exception:
            print("NSO URL is invalid. Please change in config file")
        except Exception as e:
            print(str(e))

    def get_globalrange(self, pool_name, value):
        self.globalurl_range = os.getenv(
            "NSO_URL") + "/api/running/tnt-resource-db:resource-db/global-id-pool/"
        self.localheader = {}
        self.localheader["Accept"] = 'application/vnd.yang.collection+json'
        self.localheader['Content-Type'] = 'text/plain'
        try:
            pool_name = pool_name.strip()
            item = requests.request("GET", self.globalurl_range+pool_name+"/range",
                                    headers=self.localheader, auth=(os.getenv("NSO_USERNAME"), os.getenv("NSO_PASSWORD")))
            if item.status_code == 200:
                item = item.json()
                self.range_json[pool_name] = item["collection"]["tnt-resource-db:range"]
                for each in self.range_json[pool_name]:
                    if (value >= each["min"] and value <= each["max"]):
                        return True
                return False
            else:
                return "Pool range doesn't exist"
        except requests.exceptions.InvalidURL as exception:
            print("NSO URL is invalid. Please change in config file")
        except Exception as e:
            print(str(e))

    def get_discrepancy_json(self):
        try:
            self.response = requests.request(
                "POST", self.url, headers=self.discrepancy_header, auth=(os.getenv("NSO_USERNAME"), os.getenv("NSO_PASSWORD")), data={})
            print("Discrepancy Call Completed. Getting the range for each pool")
            if self.response.status_code == 200:
                self.response = self.response.json()
                self.response = self.response["tnt-resource-db:output"]["result"]
                for each_pool in self.response:
                    pool_name = each_pool["pool"]
                    for each_vpn in each_pool["discrepancies"]:
                        consumer = each_vpn["consumer"]
                        value = each_vpn["value"]
                        existDB = each_vpn["exist-in-resource-db"]
                        flag = False
                        if pool_name in self.range_json:
                            for each in self.range_json[pool_name]:
                                if (value >= each["min"] and value <= each["max"]):
                                    flag = True
                        else:
                            if (pool_name.find(self.evi) == 0 or pool_name.find(self.bdid) == 0 or pool_name.find(self.evpl) == 0):
                                flag = self.get_localrange(pool_name, value)
                            else:
                                flag = self.get_globalrange(pool_name, value)
                        #range_pool = True if value > int(self.local_resp_json[pool_name][0]) and value < int(self.local_resp_json[pool_name][1]) else False
                        self.data.append(
                            [pool_name, consumer, value, existDB, flag])
                self.create_summary_data()
                print("Summary data generated")
                self.create_excel_discrepancy()
            else:
                print("Discrepancy Call Failed. Retry Again")
                print(self.response.reason)
        except requests.exceptions.InvalidURL as exception:
            print("NSO URL is invalid")
        except Exception as e:
            print("Exception")
            print(str(e))

    def create_excel_discrepancy(self):
        try:
            self.wb = Workbook()
            self.ws1 = self.wb.create_sheet("Summary")
            self.ws1.append(
                ["Pool Name", "True (exist-in-resource-db)", "False (exist-in-resource-db)"])
            for row in self.summary_discrepancy:
                self.ws1.append(row)

            self.ws = self.wb.create_sheet("Discrepancy Check")
            self.ws.append(["Pool Name", "Consumer Name",
                            "Value", "exist in resource DB", "exist in pool range"])
            for row in self.data:
                self.ws.append(row)

            self.wb.remove(self.wb['Sheet'])
            self.wb.save(self.excel_file)
            print("Discrepancy Excel Created")
        except Exception as e:
            print(str(e))


class disable_allocation:
    def __init__(self):
        self.disablel2_xml = os.getenv("disable_l2_xml")
        self.disablel3_xml = os.getenv("disable_l3_xml")
        self.excel_file = "NSO_disable_allocation" + dt_string + ".xlsx"
        self.url = os.getenv("NSO_URL") + "/api/query"

        self.disable_header = {}
        self.disable_header["Accept"] = 'application/vnd.yang.collection+json'
        self.disable_header['Content-Type'] = 'text/plain'
        self.disable_excel_payload = []
        self.disable_json_payload = []
        self.disable_summary = []

    def create_summary_data(self):
        l2 = set()
        l3 = set()
        for each in self.disable_excel_payload:
            if each[1] == "L3VPN":
                l3.add(each[2])
            else:
                l2.add(each[2])
        data_summary = {"L3VPN": {}, "L2VPN": {}}
        for each in l2:
            data_summary["L2VPN"][each] = [0, 0]
        for each in l3:
            data_summary["L3VPN"][each] = [0, 0]
        for each in self.disable_excel_payload:
            if each[3] == "true":
                data_summary[each[1]][each[2]
                                      ][0] = data_summary[each[1]][each[2]][0]+1
            else:
                data_summary[each[1]][each[2]
                                      ][1] = data_summary[each[1]][each[2]][1]+1
        total_false = 0
        total_true = 0
        for vpn in data_summary.keys():
            for business in data_summary[vpn].keys():
                total_true = total_true + data_summary[vpn][business][0]
                total_false = total_false + data_summary[vpn][business][1]
                self.disable_summary.append(
                    [vpn, business, data_summary[vpn][business][0], data_summary[vpn][business][1]])
        self.disable_summary.append(["", "", total_true, total_false])

    def disable_allocation_parser(self, data, type):
        for each_vpn in data:
            if type == "L3VPN":
                self.disable_excel_payload.append([each_vpn["select"][0]["value"], type, each_vpn["select"][2]["value"]+" " + each_vpn["select"]
                                                  [3]["value"], each_vpn["select"][1]["value"], each_vpn["select"][4]["value"], each_vpn["select"][5]["value"]])
            else:
                self.disable_excel_payload.append([each_vpn["select"][0]["value"], type, each_vpn["select"][2]["value"],
                                                  each_vpn["select"][1]["value"], each_vpn["select"][3]["value"], each_vpn["select"][4]["value"]])

    def get_disable_json(self):
        # L3 VPN Call Disable Allocation Call

        print()
        print("Disable Allocation Call Started")
        self.l3_response = requests.request("POST", self.url, headers=self.disable_header, auth=(
            os.getenv("NSO_USERNAME"), os.getenv("NSO_PASSWORD")), data=self.disablel3_xml)
        if self.l3_response.status_code == 200:
            self.l3_response = self.l3_response.json()
            self.l3_response = self.l3_response["tailf-rest-query:query-result"]["result"]
            self.disable_allocation_parser(self.l3_response, "L3VPN")
        else:
            print("Disable Allocation Call for L3 services Failed. Retry Again. Reason : " +
                  str(self.l3_response.reason))

        # L2 VPN Disable Allocation Call
        self.l2_response = requests.request("POST", self.url, headers=self.disable_header, auth=(
            os.getenv("NSO_USERNAME"), os.getenv("NSO_PASSWORD")), data=self.disablel2_xml)
        if self.l2_response.status_code == 200:
            self.l2_response = self.l2_response.json()
            self.l2_response = self.l2_response["tailf-rest-query:query-result"]["result"]
            self.disable_allocation_parser(self.l2_response, "L2VPN")
        else:
            print("Disable Allocation Call for L2 services Failed. Retry Again . Reason : " +
                  str(self.l2_response.reason))
        self.create_summary_data()
        self.create_excel_disable()

    def create_excel_disable(self):
        try:
            self.wb_disable = Workbook()

            self.ws_summary = self.wb_disable.create_sheet("Summary")
            self.ws_summary.append(
                ["Service Type", "Business Type", "True Disable Allocation", "False Disable Allocation"])
            for row in self.disable_summary:
                self.ws_summary.append(row)

            self.ws_disable = self.wb_disable.create_sheet(
                "NSO Disable Allocation")
            self.ws_disable.append(
                ["VPN Name", "VPN Type", "Service Type", "Value", "Running Number", "Secondary Running Number"])
            for i in self.disable_excel_payload:
                self.ws_disable.append(i)

            self.wb_disable.remove(self.wb_disable['Sheet'])
            self.wb_disable.save(self.excel_file)
            print("Disable Allocation Excel Created successfully")
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            print(str(e))
            print(exc_tb.tb_lineno)


class global_pool_resources_resourcedb:

    def __init__(self):
        self.excel_file = "NSO_global_pool_resources_resourcedb" + dt_string + ".xlsx"
        self.url = os.getenv("NSO_URL") + "/api/query"
        self.header = {}
        self.header["Accept"] = 'application/vnd.yang.collection+json'
        self.header['Content-Type'] = 'text/plain'
        self.global_list_xml = os.getenv("global_pool_list_xml")
        self.global_detail_xml = os.getenv("global_pool_detail_xml")
        self.global_dict = {}
        self.summary_data = []

    def create_summary_data(self):
        for each_pool in self.global_dict.keys():
            count_set1, count_set2, count_set3, count_set4, count_set5 = 0, 0, 0, 0, 0
            for each_vpn in self.global_dict[each_pool]:
                if len(each_vpn[2]) == 0 and len(each_vpn[3]) == 0 and len(each_vpn[4]) == 0 and len(each_vpn[5]) == 0:
                    count_set1 = count_set1 + 1
                elif len(each_vpn[3]) == 0 and len(each_vpn[5]) == 0 and ((len(each_vpn[2]) == 0 and len(each_vpn[4]) != 0) or (len(each_vpn[2]) != 0 and len(each_vpn[4]) == 0)):
                    count_set2 = count_set2 + 1
                elif (len(each_vpn[2]) != 0 and len(each_vpn[3]) != 0 and len(each_vpn[4]) == 0 and len(each_vpn[5]) == 0) or (len(each_vpn[2]) == 0 and len(each_vpn[3]) == 0 and len(each_vpn[4]) != 0 and len(each_vpn[5]) != 0):
                    count_set3 = count_set3 + 1
                elif len(each_vpn[2]) != 0 and len(each_vpn[4]) != 0 and ((len(each_vpn[3]) == 0 and len(each_vpn[5]) != 0) or (len(each_vpn[3]) != 0 and len(each_vpn[5]) == 0)):
                    count_set4 = count_set4 + 1
                elif len(each_vpn[2]) != 0 and len(each_vpn[3]) == 0 and len(each_vpn[4]) != 0 and len(each_vpn[5]) == 0:
                    count_set5 = count_set5 + 1
            self.summary_data.append(
                [each_pool, count_set1, count_set2, count_set3, count_set4, count_set5])

    def get_global_pool_list(self):
        print("Global Pool Call Started")
        try:
            self.global_resp = requests.request("POST", self.url, headers=self.header, auth=(
                os.getenv("NSO_USERNAME"), os.getenv("NSO_PASSWORD")), data=self.global_list_xml)
            if self.global_resp.status_code == 200:
                self.global_resp = self.global_resp.json()
                self.global_pool_list = []
                for pool in self.global_resp["tailf-rest-query:query-result"]["result"]:
                    self.global_pool_list.append(pool["select"][0]["value"])
                # print(self.global_pool_list)
                self.get_global_pool_details()
            else:
                print(self.global_resp.reason)
                print("Global Pool Call Failed")
        except requests.exceptions.InvalidURL as exception:
            print("NSO URL is invalid. Please change in config file")
        except Exception as e:
            print(str(e))

    def get_global_pool_details(self):
        try:
            for each in self.global_pool_list:
                each_resource = ET.fromstring(self.global_detail_xml)
                each_resource[0].text = "/tnt-resource-db:resource-db/global-id-pool[name='" + \
                    each+"']/allocation"
                each_resource = ET.tostring(each_resource)
                print(each + " Call Initiated ")
                response = requests.request(
                    "POST", self.url, headers=self.header, auth=(os.getenv("NSO_USERNAME"), os.getenv("NSO_PASSWORD")), data=each_resource)
                response = response.json()
                self.global_dict[each] = []
                if "result" in response["tailf-rest-query:query-result"]:
                    for res in response["tailf-rest-query:query-result"]["result"]:
                        self.global_dict[each].append([res["select"][0]["value"], res["select"][1]["value"], res["select"]
                                                      [2]["value"], res["select"][3]["value"], res["select"][4]["value"], res["select"][5]["value"]])
                else:
                    print("No Reservations in "+each)
            # with open("global_pool.json","w") as json_file:
                # json.dump(self.global_dict,json_file)
            self.create_summary_data()
            self.create_excel_global_pool()
        except Exception as e:
            print(str(e))

    def create_excel_global_pool(self):
        try:
            self.wb_global_pool = Workbook()
            self.ws_summary = self.wb_global_pool.create_sheet("Summary")
            self.ws_summary.append(["Pool Name", "No Allocated Values", "Value reserved(Not Used)", "Value used by service",
                                   "Primary value used by service, migration value reserved", "Both Primary and migration value reserved"])
            for row in self.summary_data:
                self.ws_summary.append(row)

            for pool in self.global_dict.keys():
                self.ws_global_pool = self.wb_global_pool.create_sheet(pool)
                self.ws_global_pool.append(["VPN Name", "native", "used-value", "used-by",
                                            "secondary-used-value", "secondary-used-by"])
                for row in self.global_dict[pool]:
                    self.ws_global_pool.append(row)

            std = self.wb_global_pool['Sheet']
            self.wb_global_pool.remove(std)
            self.wb_global_pool.save(self.excel_file)
            print("Global Pool Excel Created")
        except Exception as e:
            print(str(e))


class local_pool_resources_resource_db:
    def __init__(self):
        self.excel_file = "NSO_local_pool_resources_resource_db" + dt_string + ".xlsx"
        self.url = os.getenv("NSO_URL") + "/api/query"
        self.header = {}
        self.header["Accept"] = 'application/vnd.yang.collection+json'
        self.header['Content-Type'] = 'text/plain'
        self.local_list_xml = os.getenv("local_pool_list_xml")
        self.local_detail_xml = os.getenv("local_pool_detail_xml")
        self.local_pool_data = {}
        self.summary_data = []
        self.local_pool_json={}
        self.localurl_range_url = os.getenv("NSO_URL") + "/api/running/tnt-resource-db:resource-db/local-id-pool/?select=range;name;locality"

    def get_range(self,pool,locality,primary,secondary):
        pri_rem,sec_rem = "",""
        if pool+'.'+locality in self.local_pool_json:
            range_vpn = self.local_pool_json[pool+'.'+locality]
            if primary != "":
                pri_rem = "OUT OF RANGE"
                #print(primary,range_vpn)
                for i in range_vpn:
                    #print(range(int(i["min"]),int(i["max"])),primary in range(int(i["min"]),int(i["max"])))
                    if int(i["min"]) <= int(primary) <= int(i["max"]):
                        pri_rem = "IN RANGE"
                        break
            if secondary != "":
                sec_rem = "OUT OF RANGE"
                for i in range_vpn:
                    if int(i["min"]) <= int(secondary) <= int(i["max"]):
                        sec_rem = "IN RANGE"
                        break
        return pri_rem,sec_rem

    def get_local_pool_data(self):
        try:
            self.local_pool_json_data = requests.request("GET", self.localurl_range_url, headers=self.header, auth=(os.getenv("NSO_USERNAME"), os.getenv("NSO_PASSWORD")))
            if self.local_pool_json_data.status_code ==200:
                self.local_pool_json_data = self.local_pool_json_data.json()["collection"]["tnt-resource-db:local-id-pool"]
                for each_data in self.local_pool_json_data:
                    if 'range' in each_data:
                        self.local_pool_json[each_data["name"]+'.'+each_data["locality"]] = each_data['range']
                with open('check.json',"w") as json_file:
                    json.dump(self.local_pool_json,json_file)
            else:
                print(self.local_pool_json_data.reason)
        except Exception as e:
            print(str(e))

    def create_summary_data(self):
        for each_pool in self.local_pool_data.keys():
            count_set1, count_set2, count_set3, count_set4, count_set5 = 0, 0, 0, 0, 0
            for each_vpn in self.local_pool_data[each_pool]:
                if len(each_vpn[3]) == 0 and len(each_vpn[4]) == 0 and len(each_vpn[5]) == 0 and len(each_vpn[6]) == 0:
                    count_set1 = count_set1 + 1
                elif len(each_vpn[4]) == 0 and len(each_vpn[6]) == 0 and ((len(each_vpn[3]) == 0 and len(each_vpn[5]) != 0) or (len(each_vpn[3]) != 0 and len(each_vpn[5]) == 0)):
                    count_set2 = count_set2 + 1
                elif (len(each_vpn[3]) != 0 and len(each_vpn[4]) != 0 and len(each_vpn[5]) == 0 and len(each_vpn[6]) == 0) or (len(each_vpn[3]) == 0 and len(each_vpn[4]) == 0 and len(each_vpn[5]) != 0 and len(each_vpn[6]) != 0):
                    count_set3 = count_set3 + 1
                elif len(each_vpn[3]) != 0 and len(each_vpn[5]) != 0 and ((len(each_vpn[4]) == 0 and len(each_vpn[6]) != 0) or (len(each_vpn[4]) != 0 and len(each_vpn[6]) == 0)):
                    count_set4 = count_set4 + 1
                elif len(each_vpn[3]) != 0 and len(each_vpn[4]) == 0 and len(each_vpn[5]) != 0 and len(each_vpn[6]) == 0:
                    count_set5 = count_set5 + 1
            self.summary_data.append(
                [each_pool, count_set1, count_set2, count_set3, count_set4, count_set5])

    def get_local_pool_list(self):
        print("Local Pool Call Started")
        self.local_resp = requests.request("POST", self.url, headers=self.header, auth=(
            os.getenv("NSO_USERNAME"), os.getenv("NSO_PASSWORD")), data=self.local_list_xml)
        if self.local_resp.status_code == 200:
            self.local_resp = self.local_resp.json()
            self.local_pool_list = []
            for pool in self.local_resp["tailf-rest-query:query-result"]["result"]:
                self.local_pool_list.append(
                    (pool["select"][0]["value"], pool["select"][1]["value"]))
            # print(self.local_pool_list)
            # print("Hello")
            n = 5
            final = [self.local_pool_list[i * n:(i + 1) * n]
                     for i in range((len(self.local_pool_list) + n - 1) // n)]
            for i in range(0, len(final)):
                self.local_pool_list = final[i]
                asyncio.run(self.get_local_pool_details())
                time.sleep(1)
            self.create_summary_data()
            # with open("local.json","w") as json_file:
            # json.dump(self.local_pool_data,json_file)
            self.create_excel_local()
        else:
            print("Local Pool Call Failed")
            print(self.local_resp.reason)

    async def get_local_pool_function(self, session, each_resource, each_tuple1, each_tuple2):
        async with session.post(self.url, headers=self.header, data=each_resource, auth=aiohttp.BasicAuth(login=os.getenv("NSO_USERNAME"), password=os.getenv("NSO_PASSWORD")), ssl=False) as resp:
            data = await resp.json()
            data["tailf-rest-query:query-result"]["pool"] = each_tuple1
            data["tailf-rest-query:query-result"]["locality"] = each_tuple2
            return data

    async def get_local_pool_details(self):
        async with aiohttp.ClientSession() as session:
            tasks = []
            # print(len(self.local_pool_list))
            for each_tuple in self.local_pool_list:
                print("Getting "+each_tuple[0] +
                      " "+each_tuple[1]+" details")
                each_resource = ET.fromstring(self.local_detail_xml)
                each_resource[0].text = "/tnt-resource-db:resource-db/local-id-pool[name='" + \
                    each_tuple[0]+"'][locality='" + \
                    each_tuple[1]+"']/allocation"
                each_resource = ET.tostring(each_resource)
                tasks.append(asyncio.ensure_future(self.get_local_pool_function(
                    session, each_resource, each_tuple[0], each_tuple[1])))
            self.response = await asyncio.gather(*tasks)
            for each_item in self.response:
                if "result" in each_item["tailf-rest-query:query-result"]:
                    if each_item["tailf-rest-query:query-result"]["pool"] not in self.local_pool_data:
                        self.local_pool_data[each_item["tailf-rest-query:query-result"]["pool"]] = []
                    for record in each_item["tailf-rest-query:query-result"]["result"]:
                        primary,secondary = self.get_range(each_item["tailf-rest-query:query-result"]["pool"],each_item["tailf-rest-query:query-result"]["locality"],record["select"][2]["value"],record["select"][4]["value"])
                        self.local_pool_data[each_item["tailf-rest-query:query-result"]["pool"]].append([each_item["tailf-rest-query:query-result"]["locality"], record["select"][0]["value"], record["select"][1]["value"],
                                                                                                         record["select"][2]["value"], record["select"][3]["value"], record["select"][4]["value"], record["select"][5]["value"],primary,secondary])
        await session.close()

    def create_excel_local(self):
        try:
            self.wb_local_pool = Workbook()
            self.ws_summary = self.wb_local_pool.create_sheet("Summary")
            self.ws_summary.append(["Pool Name", "No Allocated Values", "Value reserved(Not Used)", "Value used by service",
                                   "Primary value used by service, migration value reserved", "Both Primary and migration value reserved"])
            for row in self.summary_data:
                self.ws_summary.append(row)

            for resource in self.local_pool_data.keys():
                self.ws_local_pool = self.wb_local_pool.create_sheet(resource)
                self.ws_local_pool.append(["Locality", "VPN Name", "native", "used-value",
                                           "used-by", "secondary-used-value", "secondary-used-by","Primary Allocation Remark","Secondary Allocation Remark"])
                for row in self.local_pool_data[resource]:
                    self.ws_local_pool.append(row)
            self.wb_local_pool.remove(self.wb_local_pool['Sheet'])
            self.wb_local_pool.save(self.excel_file)
            print("Local Pool Excel Created")
        except Exception as e:
            print(str(e)+"Hi")


class huawei_platform_range_verification:
    def __init__(self):
        self.huawei_range_json = "huawei_range.json"
        self.huawei_range_excel = "huawei_range_excel.xlsx"
        self.header = {}
        self.header["Accept"] = 'application/vnd.yang.collection+json'
        self.header["Content-Type"] = 'text/plain'
        self.platform_url = os.getenv("NSO_URL") + "/api/query"
        self.local_pool_range_url = os.getenv(
            "NSO_URL") + "/api/running/tnt-resource-db:resource-db/local-id-pool?select=range;name;locality"
        self.device_location_code_url = os.getenv(
            "NSO_URL") + "/api/running/iDB:inventoryDB/device-inventory?select=location-code;location-type"
        self.location_code_list = []
        self.local_pool_data = {} # Local Pool Range details
        self.excel_list = []
        self.devices = [] #inventory devices
        self.huawei_dict = {} #Huawei excel range
        self.platform_details ={} # platform details

    def get_huawei_range_json(self):
        try:
            huawei_file = open(self.huawei_range_json)
            self.huawei_dict = json.load(huawei_file)
            print("Got Huawei to-be-range")
            # print(self.huawei_dict)
        except FileNotFoundError as e:
            print(
                f"{self.huawei_range_json} file does not exist. Searching for {self.huawei_range_excel}")
            print(
                f"Excel should exist with name {self.huawei_range_excel} and have sheetname as range")
            workbook = load_workbook(filename=self.huawei_range_excel)
            sheet = workbook["range"]
            for value in sheet.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
                if value[0] not in self.huawei_dict:
                    self.huawei_dict[value[0]] = {}
                self.huawei_dict[value[0]][value[1]] = [value[2], value[3]]
            with open(self.huawei_range_json, "w") as json_file:
                json.dump(self.huawei_dict, json_file)
            print("Got Huawei to-be-range")

    def get_inventory_devices(self):
        try:
            self.device_inventory_location_code = requests.request(
                "GET", self.device_location_code_url, headers=self.header, auth=(os.getenv("NSO_USERNAME"), os.getenv("NSO_PASSWORD")))
            if self.device_inventory_location_code.status_code == 200:
                self.device_inventory_location_code = self.device_inventory_location_code.json()
                print("No of Huawei Location Code: ",len(self.device_inventory_location_code["collection"]["inventoryDB:device-inventory"] ))
                for each_location_code in self.device_inventory_location_code["collection"]["inventoryDB:device-inventory"]:
                    if each_location_code["location-type"] == "Huawei":
                        self.location_code_list.append(
                            f"http://127.0.0.1:8080/api/running/iDB:inventoryDB/device-inventory/{each_location_code['location-code']},Huawei/devices?select=device")
                #self.location_code_list = self.location_code_list[350:]
                #print(len(self.location_code_list ))
                n = 5
                final = [self.location_code_list[i * n:(i + 1) * n]
                         for i in range((len(self.location_code_list) + n - 1) // n)]
                for i in range(0, len(final)):
                    self.inventory_location_code_chunk = final[i]
                    asyncio.run(self.get_inventory_devices_async())
                    time.sleep(1)
                #print(self.devices)
                print("Number of Huawei Devices : ",len(self.devices))
                print("Got Inventory Details")
            else:
                print(self.device_inventory_location_code.reason)
                print("Here")
        except Exception as e:
            print(str(e))

    async def get_api_call_inventory(self, session, url):
        async with session.get(url, headers=self.header, auth=aiohttp.BasicAuth(login=os.getenv("NSO_USERNAME"), password=os.getenv("NSO_PASSWORD")), ssl=False) as resp:
            data = await resp.json()
            return data

    async def get_inventory_devices_async(self):
        async with aiohttp.ClientSession() as session:
            tasks = []
            for each_location_code in self.inventory_location_code_chunk:
                print(each_location_code)
                tasks.append(asyncio.ensure_future(self.get_api_call_inventory(session, each_location_code)))
            self.inventory_response = await asyncio.gather(*tasks)
            #print(self.inventory_response)
            for each_response in self.inventory_response:
                if each_response is not None:
                    for i in each_response["collection"]["inventoryDB:devices"]:
                        #print(i)
                        self.devices.append(i["device"])
        await session.close()
    
    def get_platform_details(self):
        try:
            platform_data_respose = requests.request("POST", self.platform_url, headers=self.header, auth=(
                os.getenv("NSO_USERNAME"), os.getenv("NSO_PASSWORD")), data=os.getenv("huawei_platform"))
            if platform_data_respose.status_code == 200:
                platform_data_respose = platform_data_respose.json()
                #print(platform_data_respose)
                for each in platform_data_respose["tailf-rest-query:query-result"]["result"]:
                    self.platform_details[each["select"][4]["value"]]=[each["select"][0]["value"], each["select"][1]["value"], each["select"][2]["value"], each["select"]
                                    [3]["value"], each["select"][4]["value"], each["select"][5]["value"], each["select"][6]["value"]]
                print("Got Platform Details")
                with open("platform_json.json","w") as json_file:
                    json.dump(self.platform_details,json_file)
            else:
                print(platform_data_respose.reason)
                raise Exception("Platform Data unavailable")
        except Exception as e:
            print("Get Platform data Failed")
            print(str(e))
        
    def get_local_pool_range(self):
        try:
            local_pool_data_resp = requests.request("GET", self.local_pool_range_url, headers=self.header, auth=(
                os.getenv("NSO_USERNAME"), os.getenv("NSO_PASSWORD")))
            if local_pool_data_resp.status_code == 200:
                local_pool_data_resp = local_pool_data_resp.json()
                # print(local_pool_data)
                for each in local_pool_data_resp["collection"]["tnt-resource-db:local-id-pool"]:
                    if 'range' in each:
                        self.local_pool_data[each['name'] +
                                             '.'+each['locality']] = each["range"]
                    else:
                        self.local_pool_data[each['name'] +
                                             '.'+each['locality']] = None
                print("Got Local Pool Range")
            else:
                print(local_pool_data_resp.reason)
                raise Exception("Resource DB Range Call Failed")
        except Exception as e:
            print("Local Pool Range Call failed")
            print(str(e))

    def local_range(self, pool_name, device):
        if pool_name+'.'+device in self.local_pool_data:
            range_loc = self.local_pool_data[pool_name+'.'+device]
            if range_loc is not None:
                if len(range_loc) == 1:
                    return str(range_loc[0]["min"])+"-"+str(range_loc[0]["max"]), ""
                else:
                    return str(range_loc[0]["min"])+"-"+str(range_loc[0]["max"]), "Multiple Range Present"
        return "", ""

    def huawei_data_manipulation(self):
        for each_device in self.devices:
            current_evpl_range, evpl_range_remark = self.local_range(
                "evpl_id_l2_eline_evpn", each_device)
            current_bdid_range, bdid_range_remark = self.local_range(
                "bd_id_l2_elan_etree_evpn", each_device)
            a,b = "Mismatch","Mismatch"
            if each_device in self.platform_details:
                temp_platform = self.platform_details[each_device]
                if temp_platform[3] in self.huawei_dict and temp_platform[5] in self.huawei_dict[temp_platform[3]]:
                    if current_evpl_range == self.huawei_dict[temp_platform[3]][temp_platform[5]][0]:
                        a="Match" 
                    elif current_evpl_range == "":
                        a=""
                    if current_bdid_range == self.huawei_dict[temp_platform[3]][temp_platform[5]][1]:
                        b = "Match"
                    elif current_bdid_range == "":
                        b=""
                    self.excel_list.append([temp_platform[0],temp_platform[1],temp_platform[2],temp_platform[3],each_device,temp_platform[5],temp_platform[6],current_evpl_range,self.huawei_dict[temp_platform[3]][temp_platform[5]][0],a,evpl_range_remark,current_bdid_range,self.huawei_dict[temp_platform[3]][temp_platform[5]][1],b,bdid_range_remark])
                else:
                    self.excel_list.append([temp_platform[0],temp_platform[1],temp_platform[2],temp_platform[3],each_device,temp_platform[5],temp_platform[6],current_evpl_range,"","",evpl_range_remark,current_bdid_range,"","",bdid_range_remark])
            else:
                self.excel_list.append(["","","","",each_device,"","",current_evpl_range,"","",evpl_range_remark,current_bdid_range,"","",bdid_range_remark])
        self.create_excel()
    
    def create_excel(self):
        try:
            wb = Workbook()
            excel_file = "NSO_huawei_platform_range_verification_" + dt_string + ".xlsx"
            ws = wb.create_sheet("NCEIP Details")
            ws.append(["uuid", "mgmt_ip",
                       "loopback_ip ", "detail-dev-type-name", "name", "software-version", "as-number", "current-evpl-range", "to-be-evpl-range", "evpl_id_mismatch_check", "evpl range remark", "current-bdid-range", "to-be-bdid-range", "bd_id_mismatch_check", "bdid range remark"])
            for row in self.excel_list:
                ws.append(row)
            wb.remove(wb['Sheet'])
            wb.save(excel_file)
            print("Excel Created")
        except Exception as e:
            print(str(e))


class local_pool_ranges:
    def __init__(self):
        self.excel_list=[]
        self.excel_file = "NSO_local_pool_ranges_" + dt_string + ".xlsx"
        self.localurl_range_url = os.getenv("NSO_URL") + "/api/running/tnt-resource-db:resource-db/local-id-pool/?select=range;name;locality"
        self.localheader = {}
        self.localheader["Accept"] = 'application/vnd.yang.collection+json'
        self.localheader['Content-Type'] = 'text/plain'

    def create_excel(self):
        try:
            self.wb = Workbook()
            self.ws = self.wb.create_sheet("Local Pool Range")
            self.ws.append(["Pool Name", "Device","Ranges"])
            for row in self.excel_list:
                self.ws.append(row)
            self.wb.remove(self.wb['Sheet'])
            self.wb.save(self.excel_file)
            print("Local Pool Range Excel Created")
        except Exception as e:
            print(str(e))  
    
    def get_range(self,range):
        range_list=[]
        for each_range in range:
            range_list.append(str(each_range['min']) +'-' +str(each_range['max']))
        return range_list

    def get_local_pool_data(self):
        try:
            self.local_pool_json_data = requests.request("GET", self.localurl_range_url, headers=self.localheader, auth=(os.getenv("NSO_USERNAME"), os.getenv("NSO_PASSWORD")))
            if self.local_pool_json_data.status_code ==200:
                self.local_pool_json_data = self.local_pool_json_data.json()["collection"]["tnt-resource-db:local-id-pool"]
                for each_data in self.local_pool_json_data:
                    if 'range' not in each_data:
                        self.excel_list.append([each_data["name"],each_data["locality"]])
                    else:
                        new_list = self.get_range(each_data['range'])
                        self.excel_list.append([each_data["name"],each_data["locality"]]+new_list)
                self.create_excel()

            else:
                print(self.local_pool_json_data.reason)
        except Exception as e:
            print(str(e))



if __name__ == "__main__":
    print("_______________________________________Script Starting_______________________________________")
    '''
    start_time = time.time()
    dc_obj = discrepancy_check()
    dc_obj.get_discrepancy_json()
    print("Time Taken to run discrepancy script ---> %s seconds" %
          (time.time() - start_time))
    
    start_time = time.time()
    da_obj = disable_allocation()
    da_obj.get_disable_json()
    print("Time Taken to run disable allocation script ---> %s seconds" %
          (time.time() - start_time))
    '''
    start_time = time.time()
    global_pool_obj = global_pool_resources_resourcedb()
    global_pool_obj.get_global_pool_list()
    print("Time Taken to run Global pool script ---> %s seconds" %
          (time.time() - start_time))
    
    start_time = time.time()
    local_pool_obj = local_pool_resources_resource_db()
    local_pool_obj.get_local_pool_data()
    local_pool_obj.get_local_pool_list()
    print("Time Taken to run Local pool script ---> %s seconds" %
          (time.time() - start_time))
    '''
    start_time = time.time()
    local_pool_obj = huawei_platform_range_verification()
    #local_pool_obj.get_huawei_range_json()
    local_pool_obj.get_platform_details()
    #local_pool_obj.get_inventory_devices()
    #local_pool_obj.get_local_pool_range()
    #local_pool_obj.huawei_data_manipulation()
    print("Time Taken to run Local pool script ---> %s seconds" %
          (time.time() - start_time))

    start_time = time.time()
    local_pool_obj = local_pool_ranges()
    local_pool_obj.get_local_pool_data()
    print("Time Taken to run Local pool script ---> %s seconds" %
          (time.time() - start_time))
    '''
    print("_______________________________________Script Completed_______________________________________")
